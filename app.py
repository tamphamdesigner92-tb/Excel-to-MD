import os
import io
import re
import zipfile
import uuid
from flask import Flask, request, render_template, send_file, jsonify
from openpyxl import load_workbook
from openpyxl.cell.rich_text import CellRichText, TextBlock

app = Flask(__name__)

UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

# Định nghĩa các dấu hiệu kết thúc câu để ứng dụng nhận biết
ENDING_MARKS = ('.', '!', '?', '…', ';', ':', '"', '”', "'", ')', ']', '}')
SHOP_NAME_PATTERN = re.compile(r'\[\s*tên\W*shop\W*\]|\{\s*tên\W*shop\W*\}', re.IGNORECASE)
NOTE_PATTERN = re.compile(r'\(\s*(dấu\s+[^)]*?)\s*\)', re.IGNORECASE)
PAREN_SEGMENT_PATTERN = re.compile(r'\(([^()]*)\)')
DAU_NOTE_PREFIX_PATTERN = re.compile(r'^\s*dấu\s+', re.IGNORECASE)

def is_font_bold(font):
    return bool(font and getattr(font, 'b', False))

def is_font_struck(font):
    if not font:
        return False
    return bool(
        getattr(font, 'strike', False) or
        getattr(font, 'strikethrough', False)
    )

def mark_bold_text(text):
    if not text:
        return ""
    leading_spaces = text[:len(text) - len(text.lstrip())]
    trailing_spaces = text[len(text.rstrip()):]
    core_text = text.strip()

    if not core_text:
        return text

    # Tách các cụm trong ngoặc tròn ra khỏi phần chữ in đậm còn lại, để ưu tiên
    # đánh dấu đậm+nghiêng cho riêng cụm trong ngoặc. Bỏ qua ghi chú dạng
    # "(dấu ...)" vì nó sẽ được xử lý riêng ở replace_note_markers().
    segments = []
    plain_buffer = ""
    last_end = 0

    for match in PAREN_SEGMENT_PATTERN.finditer(core_text):
        inner = match.group(1)
        if DAU_NOTE_PREFIX_PATTERN.match(inner):
            plain_buffer += core_text[last_end:match.end()]
        else:
            plain_buffer += core_text[last_end:match.start()]
            if plain_buffer.strip():
                segments.append(("plain", plain_buffer.strip()))
            plain_buffer = ""
            if inner.strip():
                segments.append(("paren", inner.strip()))
            else:
                segments.append(("plain", "()"))
        last_end = match.end()

    plain_buffer += core_text[last_end:]
    if plain_buffer.strip():
        segments.append(("plain", plain_buffer.strip()))

    if not segments or (len(segments) == 1 and segments[0][0] == "plain"):
        return f"{leading_spaces}**{core_text}**{trailing_spaces}"

    parts = [
        f"***{seg_text}***" if seg_type == "paren" else f"**{seg_text}**"
        for seg_type, seg_text in segments
    ]

    return f"{leading_spaces}{' '.join(parts)}{trailing_spaces}"

def ensure_sentence_ending(markdown_text, plain_text):
    if not markdown_text or not plain_text:
        return markdown_text

    if plain_text.rstrip().endswith(ENDING_MARKS):
        return markdown_text

    return markdown_text + "."

def replace_shop_name_placeholder(text, shop_name):
    return SHOP_NAME_PATTERN.sub(shop_name, text)

def replace_note_markers(text):
    def replace_match(match):
        note_text = match.group(1).strip()
        return f'"""{note_text}"""'

    return NOTE_PATTERN.sub(replace_match, text)

def get_cell_markdown_text(cell):
    plain_text_parts = []
    markdown_text_parts = []

    # TRƯỜNG HỢP 1: Ô có định dạng Rich Text
    if type(cell.value) == CellRichText:
        for element in cell.value:
            if isinstance(element, TextBlock):
                if is_font_struck(element.font) or not element.text:
                    continue

                plain_text_parts.append(element.text)
                if is_font_bold(element.font):
                    markdown_text_parts.append(mark_bold_text(element.text))
                else:
                    markdown_text_parts.append(element.text)
            elif isinstance(element, str):
                plain_text_parts.append(element)
                markdown_text_parts.append(element)

    # TRƯỜNG HỢP 2: Ô bình thường
    else:
        if not is_font_struck(cell.font) and cell.value is not None:
            text = str(cell.value)
            plain_text_parts.append(text)
            if is_font_bold(cell.font):
                markdown_text_parts.append(mark_bold_text(text))
            else:
                markdown_text_parts.append(text)

    plain_text = "".join(plain_text_parts).strip()
    markdown_text = "".join(markdown_text_parts).strip()

    return ensure_sentence_ending(markdown_text, plain_text) if markdown_text else ""

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload_files():
    files = request.files.getlist('files')
    files_data = []
    
    for file in files:
        if file.filename and file.filename.endswith(('.xlsx', '.xls', '.xlsm')):
            random_filename = str(uuid.uuid4()) + os.path.splitext(file.filename)[1]
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], random_filename)
            file.save(filepath)
            
            try:
                wb = load_workbook(filepath, read_only=True)
                sheets = wb.sheetnames
                wb.close()
                
                files_data.append({
                    'secure_filename': random_filename,
                    'original_name': file.filename,
                    'sheets': sheets
                })
            except Exception as e:
                print(f"Lỗi khi đọc file {file.filename}: {e}")
            
    return jsonify(files_data)

@app.route('/process', methods=['POST'])
def process_file():
    data = request.json
    filename = data['secure_filename']
    original_name = data['original_name']
    selected_sheets = data['sheets']
    base_name = os.path.splitext(original_name)[0]
    
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    memory_file = io.BytesIO()
    with zipfile.ZipFile(memory_file, 'w', zipfile.ZIP_DEFLATED) as zf:
        try:
            wb = load_workbook(filepath, data_only=True, rich_text=True)
            
            for sheet_name in selected_sheets:
                if sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    
                    brief_col_idx = None
                    for row_idx in range(1, min(6, sheet.max_row + 1)):
                        for col_idx in range(1, sheet.max_column + 1):
                            cell = sheet.cell(row=row_idx, column=col_idx)
                            
                            plain_val = ""
                            if type(cell.value) == CellRichText:
                                plain_val = "".join([e.text if isinstance(e, TextBlock) else str(e) for e in cell.value])
                            elif cell.value is not None:
                                plain_val = str(cell.value)
                                
                            if plain_val.strip().lower() == 'brief':
                                brief_col_idx = col_idx
                                break
                        if brief_col_idx is not None:
                            break
                    
                    if brief_col_idx is not None:
                        data_lines = []
                        for row_idx in range(3, sheet.max_row + 1):
                            cell = sheet.cell(row=row_idx, column=brief_col_idx)
                            final_text = get_cell_markdown_text(cell)

                            if final_text:
                                final_text = replace_shop_name_placeholder(final_text, base_name)
                                final_text = replace_note_markers(final_text)
                                data_lines.append(final_text)
                        
                        content = "\n".join(data_lines)
                    else:
                        content = "Không tìm thấy cột 'Brief' trong Tab này."
                    
                    md_filename = f"{base_name}_{sheet_name}.md"
                    zf.writestr(md_filename, content.encode('utf-8-sig'))
            
            wb.close()
            
        except Exception as e:
            print(f"Lỗi khi xử lý file {filename}: {str(e)}")
            zf.writestr(f"Error_{filename}.md", f"Lỗi tổng quát: {str(e)}".encode('utf-8'))

    memory_file.seek(0)
    zip_filename = f"{os.path.splitext(original_name)[0]}.zip"
    
    return send_file(
        memory_file,
        mimetype='application/zip',
        as_attachment=True,
        download_name=zip_filename
    )

if __name__ == '__main__':
    app.run(debug=True, port=5000)
